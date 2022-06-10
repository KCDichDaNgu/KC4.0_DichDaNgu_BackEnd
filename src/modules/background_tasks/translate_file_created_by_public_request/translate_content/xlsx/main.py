from datetime import datetime
import io
import pickle

import openpyxl
from infrastructure.configs.language import LanguageEnum
from infrastructure.configs.translation_history import TranslationHistoryStatus
from core.utils.common import chunk_arr
from docx import Document
from typing import List
from uuid import UUID
import pymongo
from infrastructure.configs.main import GlobalConfig, get_cnf, get_mongodb_instance, MAX_RETRIES
from infrastructure.configs.task import (
    TranslationTask_TranslationCompletedResultFileSchemaV1, 
    TranslationTask_NotYetTranslatedResultFileSchemaV1, 
    TranslationTaskNameEnum, 
    TranslationTaskStepEnum, 
    StepStatusEnum
)

from infrastructure.adapters.content_translator.main import ContentTranslator 

from modules.translation_request.database.translation_request.repository import TranslationRequestRepository, TranslationRequestEntity
from modules.translation_request.database.translation_request_result.repository import TranslationRequestResultRepository, TranslationRequestResultEntity
from modules.translation_request.database.translation_history.repository import TranslationHistoryRepository, TranslationHistoryEntity
from modules.system_setting.database.repository import SystemSettingRepository

import asyncio
import aiohttp

from infrastructure.adapters.logger import Logger

from core.utils.file import get_doc_paragraphs, get_full_path
from infrastructure.configs.translation_task import RESULT_FILE_STATUS, AllowedFileTranslationExtensionEnum, FileTranslationTask_NotYetTranslatedResultFileSchemaV1, FileTranslationTask_TranslatingResultFileSchemaV1, FileTranslationTask_TranslationCompletedResultFileSchemaV1, get_file_translation_file_path, get_file_translation_target_file_name
from core.utils.document import check_if_cell_is_string, check_if_paragraph_has_text, get_common_style

from core.utils.common import get_exception_log

config: GlobalConfig = get_cnf()
db_instance = get_mongodb_instance()

LIMIT_NUM_CHAR_TRANSLATE_REQUEST = 1000

translation_request_repository = TranslationRequestRepository()
translation_request_result_repository = TranslationRequestResultRepository()
transation_history_repository = TranslationHistoryRepository()
system_setting_repository = SystemSettingRepository()

contentTranslator = ContentTranslator()

logger = Logger('Task: translate_file_created_by_public_request.translate_content.xlsx')

async def read_task_result(
    tasks_result: List[TranslationRequestResultEntity], 
    tasks: List[TranslationRequestEntity],
    translations_history: List[TranslationHistoryEntity]
):
    
    valid_tasks_mapper = {}

    task_id_1 = list(map(lambda t: t.id.value, tasks))
    task_id_2 = list(map(lambda ts: ts.props.task_id.value, tasks_result))
    task_id_3 = list(map(lambda th: th.props.task_id.value, translations_history))

    intersection_tasks_id = list(set(task_id_1) & set(task_id_2) & set(task_id_3))
    
    for task_id in intersection_tasks_id:

        task = list(filter(lambda ts: ts.id.value == task_id, tasks))[0]
        task_result = list(filter(lambda ts: ts.props.task_id.value == task_id, tasks_result))[0]
        trans_history = list(filter(lambda ts: ts.props.task_id.value == task_id, translations_history))[0]

        try:
            data = await task_result.read_data_from_file()

            if data['status'] in [RESULT_FILE_STATUS['not_yet_translated'], RESULT_FILE_STATUS['translating']] and data['file_type'] == 'xlsx':

                valid_tasks_mapper[task_id] = {
                    'task_result_content': data,
                    'task_result': task_result,
                    'trans_history': trans_history,
                    'task': task
                }
        except Exception as e:
            print(e)

    valid_tasks_id = valid_tasks_mapper.keys()

    invalid_tasks = list(filter(lambda t: t.id.value not in valid_tasks_id, tasks))

    invalid_tasks_id = list(map(lambda t: t.id.value, invalid_tasks))

    invalid_tasks_mapper = {}

    for task_id in invalid_tasks_id:

        task = list(filter(lambda ts: ts.id.value == task_id, tasks))[0]
        task_result = list(filter(lambda ts: ts.props.task_id.value == task_id, tasks_result))[0]

        data = await task_result.read_data_from_file()

        trans_history = list(filter(lambda ts: ts.props.task_id.value == task_id, translations_history))[0]

        if data['file_type'] == 'xlsx':

            invalid_tasks_mapper[task_id] = {
                'task_result': task_result,
                'trans_history': trans_history,
                'task': task
            }

    return valid_tasks_mapper, invalid_tasks_mapper

async def mark_invalid_tasks(invalid_tasks_mapper):

    result = []
    
    async with db_instance.session() as session:
        async with session.start_transaction():

            update_request = []
            
            for task_id in invalid_tasks_mapper.keys():
                task_result = invalid_tasks_mapper[task_id]['task_result'],
                trans_history = invalid_tasks_mapper[task_id]['trans_history'],
                task = invalid_tasks_mapper[task_id]['task']

                if isinstance(task_result, tuple):
                    task_result = task_result[0]

                if isinstance(trans_history, tuple):
                    trans_history = trans_history[0]
                    
                update_request.append(
                    translation_request_repository.update(
                        task, 
                        dict(step_status=StepStatusEnum.cancelled.value),
                        conditions={}
                    )
                )
                
                update_request.append(
                    transation_history_repository.update(
                        trans_history, 
                        dict(
                            status=TranslationHistoryStatus.cancelled.value
                        )
                    )
                )

            result = await asyncio.gather(*update_request)

    return result

async def main():
    
    try:
        
        system_setting = await system_setting_repository.find_one({})
        
        ALLOWED_CONCURRENT_REQUEST = 1
        
        if ALLOWED_CONCURRENT_REQUEST <= 0: return 
        
        tasks = await translation_request_repository.find_many(
            params=dict(
                current_step=TranslationTaskStepEnum.translating_language.value,
                step_status={
                    '$in': [
                        StepStatusEnum.not_yet_processed.value,
                        StepStatusEnum.in_progress.value
                    ]
                }
            ),
            limit=1,
            # order_by=[('created_at', pymongo.ASCENDING)]
        ) 
            
        if not tasks or not (tasks[0].props.task_name == TranslationTaskNameEnum.public_file_translation.value and \
            tasks[0].props.current_step == TranslationTaskStepEnum.translating_language.value and \
            tasks[0].props.file_type == AllowedFileTranslationExtensionEnum.xlsx.value and \
            tasks[0].props.step_status in [StepStatusEnum.not_yet_processed.value, StepStatusEnum.in_progress.value]): return 

        logger.debug(
            msg=f'New task translate_file_created_by_public_request.translate_content.xlsx run in {datetime.now()}'
        )

        print(f'New task translate_file_created_by_public_request.translate_content.xlsx run in {datetime.now()}')
    
        tasks = await translation_request_repository.find_many(
            params=dict(
                task_name=TranslationTaskNameEnum.public_file_translation.value,
                current_step=TranslationTaskStepEnum.translating_language.value,
                file_type=AllowedFileTranslationExtensionEnum.xlsx.value,
                step_status={
                    '$in':[StepStatusEnum.not_yet_processed.value, StepStatusEnum.in_progress.value]
                },
                # expired_date={
                #     "$gt": datetime.now()
                # }
            ),
            limit=1
        )

        tasks_id = list(map(lambda task: task.id.value, tasks))

        if len(tasks_id) == 0: 
            logger.debug(
                msg=f'An task translate_file_created_by_public_request.translate_content.xlsx end in {datetime.now()}\n'
            )
            print(f'An task translate_file_created_by_public_request.translate_content.xlsx end in {datetime.now()}\n')
            return

        tasks_result_and_trans_history_req = [
            translation_request_result_repository.find_many(
                params=dict(
                    task_id={
                        '$in': list(map(lambda t: UUID(t), tasks_id))
                    },
                    step=TranslationTaskStepEnum.translating_language.value
                )
            ),
            transation_history_repository.find_many(
                params=dict(
                    task_id={
                        '$in': list(map(lambda t: UUID(t), tasks_id))
                    }
                )
            )
        ]

        tasks_result, translations_history = await asyncio.gather(*tasks_result_and_trans_history_req)
        

        valid_tasks_mapper, invalid_tasks_mapper = await read_task_result(
            tasks=tasks, 
            tasks_result=tasks_result,
            translations_history=translations_history
        )
        
        await mark_invalid_tasks(invalid_tasks_mapper)

        valid_tasks_id = list(map(lambda t: t, list(valid_tasks_mapper)))

        # valid_tasks_id = list(map(lambda t: t.id.value, tasks))
        chunked_tasks_id = list(chunk_arr(valid_tasks_id, ALLOWED_CONCURRENT_REQUEST))
        for chunk in chunked_tasks_id:
            await execute_in_batch(valid_tasks_mapper, chunk, ALLOWED_CONCURRENT_REQUEST)

    except Exception as e:
        
        error_message = get_exception_log(e)
        
        async with db_instance.session() as session:
            async with session.start_transaction():
        
                for task in tasks:
                    
                    retry = task.props.retry + 1
                    
                    changes = dict(
                        retry=retry,
                        error_message=error_message
                    )
                    
                    if retry > MAX_RETRIES: 
                        changes = dict(
                            step_status=StepStatusEnum.cancelled.value
                        )
                        
                        transation_history_repository_record = await transation_history_repository.find_one(
                            params=dict(
                                task_id=UUID(task.id.value)
                            )
                        )
                        
                        await transation_history_repository.update(transation_history_repository_record, dict(
                            status=TranslationHistoryStatus.cancelled.value
                        ))
                    
                    await translation_request_repository.update(task, changes)
        
        print(e)

    logger.debug(
        msg=f'An task translate_file_created_by_public_request.translate_content.xlsx end in {datetime.now()}\n'
    )

    print(f'An task translate_file_created_by_public_request.translate_content.xlsx end in {datetime.now()}\n')
            

async def execute_in_batch(valid_tasks_mapper, tasks_id, allowed_concurrent_request):
    loop = asyncio.get_event_loop()

    connector = aiohttp.TCPConnector(limit=allowed_concurrent_request)

    async with aiohttp.ClientSession(connector=connector, loop=loop) as session:
        api_requests = []
        for task_id in tasks_id:
            binary_progress_file_full_path = valid_tasks_mapper[task_id]['task_result_content']['binary_progress_file_full_path']

            total_paragraphs = valid_tasks_mapper[task_id]['task_result_content']['statistic']['total_paragraphs']
            processed_paragraph_index = valid_tasks_mapper[task_id]['task_result_content']['current_progress']['processed_paragraph_index']
            total_sheets = valid_tasks_mapper[task_id]['task_result_content']['statistic']['total_sheets']
            processed_sheet_index = valid_tasks_mapper[task_id]['task_result_content']['current_progress']['processed_sheet_index']
            last_row = valid_tasks_mapper[task_id]['task_result_content']['current_progress']['last_row']
            last_col = valid_tasks_mapper[task_id]['task_result_content']['current_progress']['last_col']

            source_lang = valid_tasks_mapper[task_id]['task_result_content']['source_lang']
            target_lang = valid_tasks_mapper[task_id]['task_result_content']['target_lang']

            if source_lang == target_lang:
                async with db_instance.session() as session:

                    async with session.start_transaction():
                
                        update_request = []
                        
                        task_result = valid_tasks_mapper[task_id]['task_result'],
                        trans_history = valid_tasks_mapper[task_id]['trans_history'],
                        task = valid_tasks_mapper[task_id]['task']
                        task_result_content = valid_tasks_mapper[task_id]['task_result_content']

                        original_file_name = task_result_content['original_file_full_path'].split('/')[-1]
                        original_file_ext = original_file_name.split('.')[-1]

                        target_file_name = f'{get_file_translation_target_file_name()}.{original_file_ext}'
                        target_file_path = get_file_translation_file_path(task_id, target_file_name)
                        target_file_full_path = get_full_path(target_file_path)


                        workbook = openpyxl.load_workbook(task_result_content['original_file_full_path'])

                        workbook.save(target_file_full_path)

                        new_saved_content = FileTranslationTask_TranslationCompletedResultFileSchemaV1(
                            original_file_full_path=task_result_content['original_file_full_path'],
                            binary_progress_file_full_path=task_result_content['binary_progress_file_full_path'],
                            file_type=task_result_content['file_type'],
                            statistic=dict(
                                total_paragraphs=total_paragraphs,
                                total_sheets=total_sheets,
                            ),
                            current_progress=dict(
                                processed_paragraph_index=total_paragraphs - 1,
                                processed_sheet_index=total_sheets - 1
                            ),
                            target_file_full_path=target_file_full_path,
                            source_lang=task_result_content['source_lang'],
                            target_lang=task_result_content['target_lang'],
                            task_name=TranslationTaskNameEnum.public_file_translation.value
                        )

                        if isinstance(task_result, tuple):
                            task_result = task_result[0]

                        if isinstance(trans_history, tuple):
                            trans_history = trans_history[0]

                        update_request.append(
                            translation_request_repository.update(
                                task, 
                                dict(
                                    step_status=StepStatusEnum.completed.value,
                                    current_step=TranslationTaskStepEnum.translating_language.value
                                )
                            )
                        )
                        
                        update_request.append(
                            transation_history_repository.update(
                                trans_history, 
                                dict(
                                    status= TranslationHistoryStatus.translated.value
                                )
                            )
                        )

                        update_request.append(
                            task_result.save_request_result_to_file(
                                content=new_saved_content.json()
                            )
                        )  

                await asyncio.gather(*update_request)
            
            else:
                with (open(binary_progress_file_full_path, "rb")) as openfile:
                    workbook = openpyxl.load_workbook(pickle.load(openfile))
                char_count = 0
                ws_name_list = workbook.sheetnames

                processing_sheet= processed_sheet_index + 1;
                worksheet = workbook[ws_name_list[processing_sheet]]
                
                concat_paragraphs = []
                for r in range(last_row, worksheet.max_row+1):
                    breaker = False;
                    for c in range(1, worksheet.max_column+1):                        
                        if (r == last_row and c > last_col) or (r > last_row):                            
                            cell = (worksheet.cell(r,c))
                            if check_if_cell_is_string(cell):
                                text = cell.value      
                                concat_paragraphs.append(text)
                                char_count = char_count + len(text)
                                
                                if ((char_count + len(text)) > LIMIT_NUM_CHAR_TRANSLATE_REQUEST):
                                    breaker = True 
                                    break
                    if breaker:
                        break;
                
                concat_text = " \n ".join(concat_paragraphs)

                api_requests.append(
                    contentTranslator.translate(
                        source_text=concat_text, 
                        source_lang=source_lang,
                        target_lang=target_lang,
                        session=session
                    )
                )
            
        api_results = await asyncio.gather(*api_requests)

        async with db_instance.session() as session:

            async with session.start_transaction():
                
                update_request = []
                
                for task_id, api_result in zip(tasks_id, api_results):
                    if api_result.data == '':
                        concat_translated_text = ['']
                    else:
                        concat_translated_text = api_result.data.split("\n")[:-1]

                    task_result = valid_tasks_mapper[task_id]['task_result'],
                    trans_history = valid_tasks_mapper[task_id]['trans_history'],
                    task = valid_tasks_mapper[task_id]['task']
                    task_result_content = valid_tasks_mapper[task_id]['task_result_content']

                    original_file_full_path = task_result_content['original_file_full_path']
                    binary_progress_file_full_path = task_result_content['binary_progress_file_full_path']
                    total_paragraphs = task_result_content['statistic']['total_paragraphs']
                    total_sheets = task_result_content['statistic']['total_sheets']
                    processed_sheet_index = task_result_content['current_progress']['processed_sheet_index']
                    last_row = task_result_content['current_progress']['last_row']
                    last_col = task_result_content['current_progress']['last_col']

                    with (open(binary_progress_file_full_path, "rb")) as openfile:

                        workbook = openpyxl.load_workbook(pickle.load(openfile))
                    
                    ws_name_list = workbook.sheetnames
                        
                    processing_sheet = processed_sheet_index + 1;                   
                    
                    worksheet = workbook[ws_name_list[processing_sheet]]
                    
                    cell_index = 0 

                    for r in range(last_row, worksheet.max_row + 1):
                        breaker = False
                        for c in range(1, worksheet.max_column + 1):
                            if (r == last_row and c > last_col) or (r > last_row): 
                                cell = (worksheet.cell(r,c))

                                if check_if_cell_is_string(cell):
                                    worksheet.cell(r,c).value = concat_translated_text[cell_index]
                                    
                                    if cell_index == len(concat_translated_text) - 1:
                                        breaker = True
                                        break
                                    
                                    cell_index += 1
                                    
                        last_row = r
                        last_col = c
                        if breaker:
                            break
                    
                    if (last_row < worksheet.max_row or last_col < worksheet.max_column):
                        processing_sheet = processing_sheet - 1
                    else:
                        last_row = 1
                        last_col = 0

                    with open(binary_progress_file_full_path, 'r+b') as outp:
                        new_file = io.BytesIO()
                        workbook.save(new_file)                        
                        pickle.dump(new_file, outp, pickle.HIGHEST_PROTOCOL)

                    if (processing_sheet < total_sheets - 1):                        

                        new_saved_content = FileTranslationTask_TranslatingResultFileSchemaV1(
                            original_file_full_path=original_file_full_path,
                            binary_progress_file_full_path=binary_progress_file_full_path,
                            file_type=task_result_content['file_type'],
                            statistic=dict(
                                total_sheets=total_sheets
                            ),
                            current_progress=dict(
                                processed_sheet_index=processing_sheet,
                                last_row=last_row,
                                last_col=last_col,
                            ),
                            source_lang=task_result_content['source_lang'],
                            target_lang=task_result_content['target_lang'],
                            task_name=TranslationTaskNameEnum.public_file_translation.value
                        )

                        if isinstance(task_result, tuple):
                            task_result = task_result[0]

                        if isinstance(trans_history, tuple):
                            trans_history = trans_history[0]

                        update_request.append(
                            translation_request_repository.update(
                                task, 
                                dict(
                                    step_status=StepStatusEnum.in_progress.value,
                                    current_step=TranslationTaskStepEnum.translating_language.value
                                )
                            )
                        )
                        
                        update_request.append(
                            transation_history_repository.update(
                                trans_history, 
                                dict(
                                    status= TranslationHistoryStatus.translating.value
                                )
                            )
                        )

                        update_request.append(
                            task_result.save_request_result_to_file(
                                content=new_saved_content.json()
                            )
                        )
                    else:
                        original_file_name = task_result_content['original_file_full_path'].split('/')[-1]
                        original_file_ext = task_result_content['file_type']

                        target_file_name = f'{get_file_translation_target_file_name()}.{original_file_ext}'
                        target_file_path = get_file_translation_file_path(task_id, target_file_name)
                        target_file_full_path = get_full_path(target_file_path)

                        workbook.save(target_file_full_path)

                        new_saved_content = FileTranslationTask_TranslationCompletedResultFileSchemaV1(
                            original_file_full_path=task_result_content['original_file_full_path'],
                            binary_progress_file_full_path=task_result_content['binary_progress_file_full_path'],
                            file_type=original_file_ext,
                            statistic=dict(
                                total_sheets=total_sheets,
                            ),
                            current_progress=dict(
                                processed_sheet_index=processing_sheet,
                                last_row=last_row,
                                last_col=last_col,
                            ),
                            target_file_full_path=target_file_full_path,
                            source_lang=task_result_content['source_lang'],
                            target_lang=task_result_content['target_lang'],
                            task_name=TranslationTaskNameEnum.public_file_translation.value
                        )

                        if isinstance(task_result, tuple):
                            task_result = task_result[0]

                        if isinstance(trans_history, tuple):
                            trans_history = trans_history[0]

                        update_request.append(
                            translation_request_repository.update(
                                task, 
                                dict(
                                    step_status=StepStatusEnum.completed.value,
                                    current_step=TranslationTaskStepEnum.translating_language.value
                                )
                            )
                        )
                        
                        update_request.append(
                            transation_history_repository.update(
                                trans_history, 
                                dict(
                                    status= TranslationHistoryStatus.translated.value
                                )
                            )
                        )

                        update_request.append(
                            task_result.save_request_result_to_file(
                                content=new_saved_content.json()
                            )
                        )  

                await asyncio.gather(*update_request)
